import os
#import xml.etree.ElementTree as ET
from ultralytics import YOLO
import pandas as pd
import cv2
import numpy as np
from typing import Tuple, List, Optional
from tqdm import tqdm  # Thêm thư viện tqdm để hiển thị tiến trình
import openpyxl
from colorama import init, Fore, Style
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
#from openpyxl.utils import get_column_letter # Helper để chuyển đổi chỉ số cột sang chữ cái
init() # Khởi tạo colorama

luu_anh_cat = True
luu_anh_goc = True
luu_anh_cham = True
luu_anh_contours = True

def chen_thong_tin_len_anh(anh, thong_tin):
    # Nội dung chữ
    chu1 = f"So bao danh: {thong_tin[0]}"
    chu2 = f"Ho ten: {thong_tin[1]}"
    chu3 = f"Ma de: {thong_tin[2]}"
    chu4 = f"Diem: {thong_tin[3]}"
    chu5 = f"{thong_tin[4]}"
    # Cài đặt font
    font = cv2.FONT_HERSHEY_SIMPLEX
    co_chu = 0.4           # kích thước font (scale)
    do_day = 1           # độ dày nét
    mau_chu = (0, 0, 92)  # màu

    # Tính kích thước chữ
    #(text_w, text_h), _ = cv2.getTextSize(chu1, font, co_chu, do_day)

    # Tính vị trí để chữ nằm giữa ảnh
    vi_tri1 = (20,20)
    vi_tri2 = (20,40)
    vi_tri3 = (20,60)
    vi_tri4 = (20,80)
    vi_tri5 = (220,20)

    # Vẽ chữ lên ảnh
    cv2.putText(anh, chu1, vi_tri1, font, co_chu, mau_chu, do_day, cv2.LINE_AA)
    cv2.putText(anh, chu2, vi_tri2, font, co_chu, mau_chu, do_day, cv2.LINE_AA)
    cv2.putText(anh, chu3, vi_tri3, font, co_chu, mau_chu, do_day, cv2.LINE_AA)
    cv2.putText(anh, chu4, vi_tri4, font, co_chu, mau_chu, do_day, cv2.LINE_AA)
    cv2.putText(anh, chu5, vi_tri5, font, co_chu, mau_chu, do_day, cv2.LINE_AA)

    # Hiển thị (tùy chọn)
    # cv2.imshow("Ảnh với chữ", img)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    # return img

def tao_hinh_trong():
    # Tạo ảnh BGR đơn sắc
    h, w = 510, 630
    mau_bgr = (214, 227, 251)  # Vàng
    img = np.full((h, w, 3), mau_bgr, dtype=np.uint8)

    # Nội dung chữ
    Text_1 = "No data !"

    # Cài đặt font
    font = cv2.FONT_HERSHEY_SIMPLEX
    co_chu = 2           # kích thước font (scale)
    do_day = 4           # độ dày nét
    mau_chu = (0, 0, 192)  # màu

    # Tính kích thước chữ
    (text_w, text_h), _ = cv2.getTextSize(Text_1, font, co_chu, do_day)

    # Tính vị trí để chữ nằm giữa ảnh
    vi_tri = ((w - text_w) // 2, (h + text_h) // 2)

    # Vẽ chữ lên ảnh
    cv2.putText(img, Text_1, vi_tri, font, co_chu, mau_chu, do_day, cv2.LINE_AA)

    # Hiển thị (tùy chọn)
    # cv2.imshow("Ảnh với chữ", img)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    return img

def hien_thi_thoi_gian():
    thoi_gian_hien_tai = datetime.now()
    dinh_dang = thoi_gian_hien_tai.strftime("%Y_%m_%d - %H : %M : %S")
    return dinh_dang

def ReSize(img, ChieuCao):
    """
    Thay đổi kích thước ảnh theo chiều cao mong muốn.
    Args:
        img (numpy.ndarray): Ảnh đầu vào.
        ChieuCao (int): Chiều cao mong muốn của ảnh. Nếu là None, mặc định là 500.
    Returns:
        numpy.ndarray: Ảnh đã được thay đổi kích thước.
    """
    if ChieuCao is not None:
        target_height = ChieuCao
    else:
        target_height = 500
    original_height, original_width = img.shape[0], img.shape[1]
    if original_height == 0:
        print("Lỗi: Chiều cao ảnh gốc không hợp lệ (bằng 0).")
        return img
    scale = target_height / original_height
    new_width = int(original_width * scale)
    new_size = (new_width, target_height)
    interpolation_method = cv2.INTER_AREA if scale < 1 else cv2.INTER_LINEAR
    resized_img = cv2.resize(img, new_size, interpolation=interpolation_method)
    return resized_img

def tao_so_luong_cau_hoi(so_luong):
    if so_luong > 0:
        cau_hoi = []
        so_luong = round(so_luong)
        for i in range(1, so_luong + 1):
            cau_hoi.append(i)
        return cau_hoi

def to_mau_dap_an(anh_cham, ma_tran_dap_an, ket_qua_dap_an, dap_an_chon, dap_an_dung, to_thong_tin):
    """
    Tô màu các ô đáp án: xanh cho đúng, đỏ cho sai
    
    Args:
        anh_cham: Ảnh để vẽ lên
        ma_tran_dap_an: Ma trận tọa độ các ô đáp án
        ket_qua_dap_an: Ma trận kết quả tô của các ô
        dap_an_chon: Danh sách đáp án học sinh chọn
        dap_an_dung: Danh sách đáp án đúng
    """
    # Tô màu cho thông tin (SBD, mã đề) - giữ nguyên màu xanh
    for i in to_thong_tin:
        cv2.circle(anh_cham, i, 10, (98,250,202), 2)
    
    # Tô màu cho đáp án dựa trên đúng/sai
    for hang_idx, hang in enumerate(ma_tran_dap_an):
        if hang_idx < len(dap_an_chon) and hang_idx < len(dap_an_dung):
            dap_an_hoc_sinh = dap_an_chon[hang_idx]
            dap_an_chuan = dap_an_dung[hang_idx]
            
            for cot_idx, (x, y) in enumerate(hang):
                # Kiểm tra xem ô này có được tô hay không
                if (hang_idx < len(ket_qua_dap_an) and 
                    cot_idx < len(ket_qua_dap_an[hang_idx]) and 
                    ket_qua_dap_an[hang_idx][cot_idx]):
                    
                    # Xác định màu dựa trên đáp án đúng/sai
                    if dap_an_hoc_sinh == dap_an_chuan:
                        # Đáp án đúng - tô màu xanh lá
                        mau = (0, 255, 0)  # BGR: xanh lá
                    else:
                        # Đáp án sai - tô màu đỏ
                        mau = (0, 0, 255)  # BGR: đỏ
                    
                    cv2.circle(anh_cham, (x, y), 10, mau, 2)

def cham_bai(file_mo_hinh, thu_muc_anh, file_danh_sach, file_dap_an, output_dir, kich_thuoc_cat):
    anh_debug = None
    sbd = None
    ma_de = None
    dap_an_chon = None
    to_thong_tin = None
    to_dap_an = None
    bo_qua_anh = True

    # Kiểm tra mô hình
    model_path = os.path.join(file_mo_hinh) # , 'train', 'weights', 'best.pt')
    if not os.path.exists(model_path):
        print(f"{Fore.RED}(!) Lỗi: Không tìm thấy mô hình tại {model_path}.")
        print(f"{Fore.YELLOW}  --> Hãy nhập mô hình vào {model_path} !{Fore.RESET}")
        return

    # Kiểm tra thư mục ảnh
    if not os.path.exists(thu_muc_anh) or not os.listdir(thu_muc_anh):
        print(f"{Fore.RED}(!) Lỗi: Không tìm thấy thư mục ảnh hoặc thư mục trống tại {thu_muc_anh}{Fore.RESET}")
        return

    # Tạo thư mục gốc và các thư mục con
    ket_qua_dir = output_dir
    anh_goc_dir = os.path.join(ket_qua_dir, 'anh_goc_danh_dau')
    anh_cat_dir = os.path.join(ket_qua_dir, 'anh_da_cat')
    output_contours_dir = os.path.join(ket_qua_dir, 'anh_contours')
    anh_cham_dir = os.path.join(ket_qua_dir, 'anh_da_cham')
    excel_dir = os.path.join(ket_qua_dir, 'file_excel')
    
    os.makedirs(ket_qua_dir, exist_ok=True)
    os.makedirs(anh_goc_dir, exist_ok=True)
    os.makedirs(anh_cat_dir, exist_ok=True)
    os.makedirs(output_contours_dir, exist_ok=True)
    os.makedirs(anh_cham_dir, exist_ok=True)
    os.makedirs(excel_dir, exist_ok=True)

    print(f"\n{Fore.YELLOW}" + "-" * 20 + f"{Fore.RESET}")

    # Lặp qua tất cả các file ảnh trong thư mục
    for ten_file in tqdm(os.listdir(thu_muc_anh), desc="Đang xử lý ảnh"):
        if ten_file.endswith(('.jpg', '.jpeg', '.png')):
            duong_dan_anh_thu = os.path.join(thu_muc_anh, ten_file)
            output_goc_path = os.path.join(anh_goc_dir, ten_file)
            output_cat_path = os.path.join(anh_cat_dir, ten_file)
            output_contours_path = os.path.join(output_contours_dir, ten_file)
            output_cham_path = os.path.join(anh_cham_dir, ten_file)

            # Thực hiện dự đoán
            try:
                model = YOLO(model_path)
                results = model.predict(duong_dan_anh_thu, save=False)
                result = results[0]

                boxes = result.boxes.xyxy
                classes = result.boxes.cls
                scores = result.boxes.conf
                print(f"Đã xử lý ảnh: '{Fore.BLUE}{ten_file}{Fore.RESET}' - Đã phát hiện {Fore.YELLOW}{len(boxes)}{Fore.RESET} đối tượng:")
                toa_do_all = []
                for box, cls, score in zip(boxes, classes, scores):
                    xmin, ymin, xmax, ymax = [int(coord) for coord in box.tolist()]
                    center_x = round((xmin + xmax) / 2)
                    center_y = round((ymin + ymax) / 2)
                    toa_do_all.append((center_x, center_y))
                    print(f"  Box: {Fore.YELLOW}[{xmin}, {ymin}, {xmax}, {ymax}], Class: {int(cls)}, Score: {score:.2f}, Tâm: ({center_x}, {center_y}){Fore.RESET}")

                # Đọc ảnh bằng OpenCV để vẽ và cắt
                anh_goc = cv2.imread(duong_dan_anh_thu)
                anh_danh_dau_goc = result.plot()

                # Cắt ảnh và lưu
                if len(toa_do_all) == 4:
                    anh_da_cat = cat_anh_co_dinh(anh_goc.copy(), toa_do_all, kich_thuoc_cat)
                    if anh_da_cat is not None:
                        bo_qua_anh = False
                        if 'anh_da_cat' in locals():
                            anh_debug, anh_contours, sbd, ma_de, dap_an_chon, to_thong_tin, to_dap_an, ma_tran_dap_an, ket_qua_dap_an = nhan_dien_trac_nghiem(ten_file, anh_da_cat)
                            print(f"Đang chấm trên ảnh: {Fore.BLUE}{ten_file}{Fore.RESET}")
                            print(f"Số báo danh: {sbd}")
                            print(f"Mã đề: {ma_de}")
                            print(f"Đáp án: {dap_an_chon}")
                    else:
                        print(f"{Fore.RED}(!) CẢNH BÁO: Không thể cắt ảnh do không xác định được 4 góc hợp lệ, hãy kiểm tra lại hình {ten_file}.")
                        print(f"(!) BỎ QUA ẢNH NÀY !{Fore.RESET}")
                        bo_qua_anh = True
                else:
                    print(f"{Fore.RED}(!) CẢNH BÁO: Phát hiện {len(toa_do_all)} đối tượng, hãy kiểm tra lại hình {Fore.BLUE}'{ten_file}'{Fore.RED} hoặc nâng cấp mô hình!")
                    print(f"(!) --> BỎ QUA ẢNH NÀY !{Fore.RESET}")
                    bo_qua_anh = True
                    
                if not bo_qua_anh:
                    dap_an = Nhap_file_dap_an(file_dap_an)
                    thong_tin = Tim_thong_tin(file_danh_sach, sbd)
                    bai_thi = [ma_de, sbd, tao_so_luong_cau_hoi(10), dap_an_chon]
                    ket_qua = Cham_diem(bai_thi, dap_an, thong_tin)
                    file_ket_qua_da_tao = tao_file_excel(sbd, ket_qua, excel_dir)  # Lưu vào thư mục file_excel

                    anh_cham = anh_da_cat.copy()
                    for i in to_thong_tin:
                        cv2.circle(anh_cham, i, 10, (17,255,255), 2)
                    # Tô màu đáp án theo đúng/sai
                    to_mau_dap_an(anh_cham, ma_tran_dap_an, ket_qua_dap_an, dap_an_chon, dap_an[2], to_dap_an)

                    print(f"Số đáp án làm đúng: {ket_qua[5]}, điểm: {ket_qua[6]}")
                    # Lưu ảnh đã đánh dấu
                    if luu_anh_goc:
                        cv2.imwrite(output_goc_path, anh_danh_dau_goc)
                        print(f"{Fore.GREEN}--> Ảnh đã đánh dấu được lưu tại: {Fore.CYAN}{output_goc_path}{Fore.RESET}")
                        
                    if luu_anh_cat:
                        cv2.imwrite(output_cat_path, anh_da_cat)
                        print(f"{Fore.GREEN}--> Đã lưu ảnh cắt tại: {Fore.CYAN}{output_cat_path}{Fore.GREEN} với kích thước {Fore.YELLOW}{kich_thuoc_cat}{Fore.RESET}")
                    
                    if luu_anh_contours:
                        cv2.imwrite(output_cat_path, anh_contours)
                        print(f"{Fore.GREEN}--> Đã lưu ảnh contours tại: {Fore.CYAN}{output_contours_path}{Fore.GREEN} với kích thước {Fore.YELLOW}{kich_thuoc_cat}{Fore.RESET}")
                    
                    chen_thong_tin_len_anh(anh_cham, (sbd, ket_qua[0][4], ma_de, ket_qua[6], hien_thi_thoi_gian()))
                    if luu_anh_cham:
                        cv2.imwrite(output_cham_path, anh_cham)
                        print(f"{Fore.GREEN}--> Đã lưu ảnh cắt chấm tại: {Fore.CYAN}{output_cham_path}{Fore.GREEN} với kích thước {Fore.YELLOW}{kich_thuoc_cat}{Fore.RESET}")

                    if file_ket_qua_da_tao:
                        print(f"{Fore.GREEN}--> Đã lưu file EXCEL tại: {Fore.CYAN}{file_ket_qua_da_tao}{Style.RESET_ALL}")
                    else:
                        print(f"{Fore.RED}Quá trình gặp lỗi, không tạo được file kết quả.{Style.RESET_ALL}")

                print("-" * 50 + "\n")
                anh_danh_dau_goc_copy = ReSize(anh_danh_dau_goc, 500)
                anh_da_cat_copy = ReSize(anh_da_cat, 300)
                anh_contours_copy = ReSize(anh_contours, 300)
                anh_debug_copy = ReSize(anh_debug, 300)
                anh_cham_copy = ReSize(anh_cham, 300)

                cv2.imshow('anh danh dau goc', anh_danh_dau_goc_copy)
                cv2.imshow('Anh cat', anh_da_cat_copy)
                cv2.imshow('anh_contours', anh_contours_copy)
                cv2.imshow("Nhan dien", anh_debug_copy)
                cv2.imshow('anh ket qua cuoi cung', anh_cham_copy)
                
                anh_da_cat = tao_hinh_trong()
                anh_contours = tao_hinh_trong()
                anh_debug = tao_hinh_trong()
                anh_cham = tao_hinh_trong()

                #cv2.waitKey()
            except Exception as e:
                print(f"Lỗi khi dự đoán ảnh đã cắt: {str(e)}")
        cv2.waitKey()
    print(f"\n{Fore.YELLOW}=== ĐÃ DUYỆT HẾT ẢNH TRONG THƯ MỤC ==={Fore.RESET}\n")
          
def cat_anh_co_dinh(img, toa_do_4_diem, kich_thuoc_out):
    """
    Nắn phẳng phối cảnh và cắt ảnh với kích thước cố định.
    Args:
        img (numpy.ndarray): Ảnh đầu vào.
        toa_do_4_diem (list): List chứa 4 tuple tọa độ (x, y) của 4 góc.
        kich_thuoc_out (tuple): Tuple (width, height) chỉ định kích thước đầu ra.
    Returns:
        numpy.ndarray: Ảnh đã được nắn phẳng và cắt với kích thước cố định,
                       hoặc None nếu không đủ 4 điểm.
    """
    if toa_do_4_diem is not None and len(toa_do_4_diem) == 4:
        points = np.array(toa_do_4_diem, dtype=np.float32)
        # Sắp xếp các điểm theo thứ tự: top-left, top-right, bottom-right, bottom-left
        s = points.sum(axis=1)
        diff = np.diff(points, axis=1)
        tl = points[np.argmin(s)]
        br = points[np.argmax(s)]
        tr = points[np.argmin(diff)]
        bl = points[np.argmax(diff)]
        pts1 = np.float32([tl, tr, br, bl])

        width, height = kich_thuoc_out
        pts2 = np.float32([[0, 0], [width - 1, 0], [width - 1, height - 1], [0, height - 1]])

        matrix = cv2.getPerspectiveTransform(pts1, pts2)
        warped = cv2.warpPerspective(img, matrix, (width, height))
        print(f"Đã nắn phẳng phối cảnh và cắt ảnh với kích thước: {kich_thuoc_out}")
        return warped
    else:
        print("Không đủ 4 điểm để nắn phẳng phối cảnh.")
        return None

def nhan_dien_trac_nghiem(ten_file, anh_da_cat: str) -> Tuple[Optional[np.ndarray], str, str, List[str]]:
    """
    Tự động chấm điểm bài trắc nghiệm từ ảnh và đọc thông tin.

    Args:
        duong_dan_anh: Đường dẫn đến file ảnh bài thi.

    Returns:
        Tuple chứa (ảnh debug đã vẽ kết quả, số báo danh, mã đề, danh sách đáp án).
        Trả về (None, "", "", []) nếu có lỗi xảy ra.
    """
    try:
        anh = anh_da_cat
        # Chuyển ảnh sang thang xám và nhị phân
        anh_xam = cv2.cvtColor(anh, cv2.COLOR_BGR2GRAY)
        anh_nhi_phan = cv2.adaptiveThreshold(
            anh_xam, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 11, 3
        )

        # Phát hiện các bubble
        contours, _ = cv2.findContours(anh_nhi_phan, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        anh_contours = anh.copy()
        cv2.drawContours(anh_contours, contours, -1, (0, 255, 255), 2) # Vẽ tất cả contours màu vàng
        cac_bubble = [
            (int(x), int(y))
            for contour in contours
            for (x, y), ban_kinh in [cv2.minEnclosingCircle(contour)]
            if 8 < ban_kinh < 25 and 0.5 < cv2.contourArea(contour) / (np.pi * ban_kinh**2 + 1e-5) < 1.5
        ]

        if not cac_bubble:
            print(f"Lỗi: Không tìm thấy bubble nào trong ảnh đã cắt")
            return anh.copy(), "", "", []

        # Phân nhóm bubble theo trục x
        toa_do = np.array(cac_bubble)
        x_gia_tri = toa_do[:, 0]
        x_min, x_max = x_gia_tri.min(), x_gia_tri.max()
        nguong_sbd = x_min + (x_max - x_min) * 0.3
        nguong_ma_de = x_min + (x_max - x_min) * 0.5

        nhom_sbd = [(x, y) for x, y in cac_bubble if x < nguong_sbd]
        nhom_ma_de = [(x, y) for x, y in cac_bubble if nguong_sbd <= x < nguong_ma_de]
        nhom_dap_an = [(x, y) for x, y in cac_bubble if x >= nguong_ma_de]

        # Sắp xếp bubble thành ma trận
        def sap_xep_bubble(bubbles: List[Tuple[int, int]], so_hang: int = 10, so_cot: int = 4) -> List[List[Tuple[int, int]]]:
            if not bubbles:
                return [[] for _ in range(so_hang)]
            
            bubbles = sorted(bubbles, key=lambda b: b[1])
            khoang_cach = (bubbles[-1][1] - bubbles[0][1]) / (so_hang - 1) if so_hang > 1 else 0
            hang = [[] for _ in range(so_hang)]
            
            for x, y in bubbles:
                chi_so_hang = min(so_hang - 1, max(0, int(round((y - bubbles[0][1]) / khoang_cach)) if khoang_cach else 0))
                hang[chi_so_hang].append((x, y))
            
            for i in range(so_hang):
                hang[i].sort(key=lambda b: b[0])
            
            return hang

        ma_tran_sbd = sap_xep_bubble(nhom_sbd, 10, 4)
        ma_tran_ma_de = sap_xep_bubble(nhom_ma_de, 10, 2)
        ma_tran_dap_an = sap_xep_bubble(nhom_dap_an, 10, 4)
        to_thong_tin = []
        to_dap_an = []

        # Đọc trạng thái tô của bubble
        def doc_to(anh_nhi_phan: np.ndarray, ma_tran: List[List[Tuple[int, int]]], ban_kinh: int = 10, nguong: float = 0.5) -> List[List[bool]]:
            ket_qua = []
            for hang in ma_tran:
                dong = []
                for x, y in hang:
                    vung = anh_nhi_phan[y - ban_kinh:y + ban_kinh, x - ban_kinh:x + ban_kinh]
                    ty_le = cv2.countNonZero(vung) / (vung.size + 1e-5) if vung.size else 0
                    dong.append(ty_le > nguong)
                ket_qua.append(dong)
            return ket_qua

        ket_qua_sbd = doc_to(anh_nhi_phan, ma_tran_sbd)
        ket_qua_ma_de = doc_to(anh_nhi_phan, ma_tran_ma_de)
        ket_qua_dap_an = doc_to(anh_nhi_phan, ma_tran_dap_an)

        # Chuyển đổi kết quả thành định dạng yêu cầu
        def doc_so_bao_danh(ma_tran_kq: List[List[bool]]) -> str:
            sbd = ""
            for cot in range(min(4, len(ma_tran_kq[0]) if ma_tran_kq else 0)):
                gia_tri = "_"
                for hang in range(len(ma_tran_kq)):
                    if ma_tran_kq[hang][cot]:
                        gia_tri = str(hang)
                        break
                sbd += gia_tri
            return sbd

        def doc_ma_de(ma_tran_kq: List[List[bool]]) -> str:
            ma_de = ""
            for cot in range(min(2, len(ma_tran_kq[0]) if ma_tran_kq else 0)):
                gia_tri = "_"
                for hang in range(len(ma_tran_kq)):
                    if ma_tran_kq[hang][cot]:
                        gia_tri = str(hang)
                        break
                ma_de += gia_tri
            return ma_de

        def doc_dap_an(ma_tran_kq: List[List[bool]]) -> List[str]:
            dap_an = []
            for hang in ma_tran_kq:
                lua_chon = ""
                if len(hang) == 4:
                    if hang[0]: lua_chon += "A"
                    if hang[1]: lua_chon += "B"
                    if hang[2]: lua_chon += "C"
                    if hang[3]: lua_chon += "D"
                dap_an.append(lua_chon or "_")
            return dap_an

        so_bao_danh = doc_so_bao_danh(ket_qua_sbd)
        ma_de = doc_ma_de(ket_qua_ma_de)
        dap_an = doc_dap_an(ket_qua_dap_an)

        # Vẽ kết quả lên ảnh debug
        def ve_bubble(anh: np.ndarray, ma_tran: List[List[Tuple[int, int]]], ket_qua: List[List[bool]], 
                      mau_to: Tuple[int, int, int] = (0, 0, 255), mau_khong_to: Tuple[int, int, int] = (69, 255, 255)):
            for i, hang in enumerate(ma_tran):
                for j, (x, y) in enumerate(hang):
                    if i < len(ket_qua) and j < len(ket_qua[i]):
                        mau = mau_to if ket_qua[i][j] else mau_khong_to
                        cv2.circle(anh, (x, y), 10, mau, 2)

        def lay_danh_sach_to_thong_tin_SBD(ma_tran_sbd: List[List[Tuple[int, int]]], ket_qua_sbd: List[List[bool]]):
            for i, hang in enumerate(ma_tran_sbd):
                for j, (x, y) in enumerate(hang):
                    if i < len(ket_qua_sbd) and j < len(ket_qua_sbd[i]):
                        if ket_qua_sbd[i][j]:
                            to_thong_tin.append((x, y))
        def lay_danh_sach_to_thong_tin_ma_de(ma_tran_ma_de: List[List[Tuple[int, int]]], ket_qua_ma_de: List[List[bool]]):
            for i, hang in enumerate(ma_tran_ma_de):
                for j, (x, y) in enumerate(hang):
                    if i < len(ket_qua_ma_de) and j < len(ket_qua_ma_de[i]):
                        if ket_qua_ma_de[i][j]:
                            to_thong_tin.append((x, y))
        def lay_danh_sach_to_dap_an(ma_tran_dap_an: List[List[Tuple[int, int]]], ket_qua_dap_an: List[List[bool]]):
            for i, hang in enumerate(ma_tran_dap_an):
                for j, (x, y) in enumerate(hang):
                    if i < len(ket_qua_dap_an) and j < len(ket_qua_dap_an[i]):
                        if ket_qua_dap_an[i][j]:
                            to_dap_an.append((x, y))


        anh_debug = anh.copy()
        ve_bubble(anh_debug, ma_tran_sbd, ket_qua_sbd)
        ve_bubble(anh_debug, ma_tran_ma_de, ket_qua_ma_de)
        ve_bubble(anh_debug, ma_tran_dap_an, ket_qua_dap_an)
        lay_danh_sach_to_thong_tin_SBD(ma_tran_sbd, ket_qua_sbd)
        lay_danh_sach_to_thong_tin_ma_de(ma_tran_ma_de, ket_qua_ma_de)
        lay_danh_sach_to_dap_an(ma_tran_dap_an, ket_qua_dap_an)

        if len(nhom_sbd) != 40:
            print(f"{Fore.RED}(!) CẢNH BÁO: các ô ở vùng SỐ BÁO DANH không được phát hiện đầy đủ, ảnh hưởng đến độ chính xác !")
            print(f"{Fore.YELLOW}--> Hãy thử chụp lại {Fore.BLUE}{ten_file}!{Fore.RESET}")
        if len(nhom_ma_de) != 20:
            print(f"{Fore.RED}(!) CẢNH BÁO: các ô ở vùng MÃ ĐỀ không được phát hiện đầy đủ, ảnh hưởng đến độ chính xác !")
            print(f"{Fore.YELLOW}--> Hãy thử chụp lại {Fore.BLUE}{ten_file}!{Fore.RESET}")
        if len(nhom_dap_an) != 40:
            print(f"{Fore.RED}(!) CẢNH BÁO: các ô ở vùng ĐÁP ÁN không được phát hiện đầy đủ, ảnh hưởng đến độ chính xác !")
            print(f"{Fore.YELLOW}--> Hãy thử chụp lại {Fore.BLUE}{ten_file}!{Fore.RESET}")

        matran = [ma_tran_sbd, ma_tran_ma_de, ma_tran_dap_an]
        return anh_debug, anh_contours, so_bao_danh, ma_de, dap_an, to_thong_tin, to_dap_an, ma_tran_dap_an, ket_qua_dap_an

    except Exception as e:
        print(f"Lỗi không xác định: {e}")
        return None, "", "", []

def Nhap_file_dap_an(file_DA):
    # Đọc file đáp án, không có header để chỉ số cột là số nguyên
    #print("Đang nhập file đáp án...")
    df_da = pd.read_excel(file_DA, header=None)
    mon_thi = df_da.iloc[df_da.iloc[0:df_da.last_valid_index(), 0][(df_da.iloc[0:df_da.last_valid_index(), 0] == "Môn thi")].index,1].item()
    ma_de = df_da.iloc[df_da.iloc[0:df_da.last_valid_index(), 0][(df_da.iloc[0:df_da.last_valid_index(), 0] == "Mã đề")].index,1].item()
    tong_so_cau = df_da.iloc[df_da.iloc[0:df_da.last_valid_index(), 0][(df_da.iloc[0:df_da.last_valid_index(), 0] == "Tổng số câu")].index,1].item()
    dap_an = [[mon_thi, ma_de, tong_so_cau], [], [], []]
    
    for i in range(tong_so_cau):
        cau_hoi_bai_thi_value = df_da.iloc[df_da.iloc[0:df_da.last_valid_index(), 0][(df_da.iloc[0:df_da.last_valid_index(), 0] == "Câu hỏi")].index + 1 + i,0].item()
        dap_an_dung_value = df_da.iloc[df_da.iloc[0:df_da.last_valid_index(), 1][(df_da.iloc[0:df_da.last_valid_index(), 1] == "Đáp án")].index + 1 + i,1].item()
        muc_diem_value = df_da.iloc[df_da.iloc[0:df_da.last_valid_index(), 2][(df_da.iloc[0:df_da.last_valid_index(), 2] == "Mức điểm")].index + 1 + i,2].item()        
        dap_an[1].append(cau_hoi_bai_thi_value) # Thêm giá trị cột 0 vào list đầu tiên
        dap_an[2].append(dap_an_dung_value)    # Thêm giá trị cột 1 vào list thứ hai
        dap_an[3].append(muc_diem_value)      # Thêm giá trị cột 2 vào list thứ ba
    print("Nhập file đáp án xong !")
    return dap_an

def Tim_thong_tin(file_danh_sach, SBD_can_tim):
    df_ds = pd.read_excel(file_danh_sach, header=None)
    # lấy thông tin
    infor_ds = df_ds.iloc[0:4,0:2]
    if "_" not in SBD_can_tim and SBD_can_tim in df_ds.iloc[0:df_ds.last_valid_index(), 1]:
        truong = df_ds.iloc[df_ds.iloc[0:df_ds.last_valid_index(), 0][(df_ds.iloc[0:df_ds.last_valid_index(), 0] == "TRƯỜNG")].index,1].item()
        khoa = df_ds.iloc[df_ds.iloc[0:df_ds.last_valid_index(), 0][(df_ds.iloc[0:df_ds.last_valid_index(), 0] == "KHOA")].index,1].item()
        nganh = df_ds.iloc[df_ds.iloc[0:df_ds.last_valid_index(), 0][(df_ds.iloc[0:df_ds.last_valid_index(), 0] == "NGÀNH")].index,1].item()
        lop = df_ds.iloc[df_ds.iloc[0:df_ds.last_valid_index(), 0][(df_ds.iloc[0:df_ds.last_valid_index(), 0] == "LỚP")].index,1].item()
        ten = df_ds.iloc[df_ds.iloc[0:df_ds.last_valid_index(), 1][(df_ds.iloc[0:df_ds.last_valid_index(), 1] == SBD_can_tim)].index,0].item()
    else:
        truong = "None"
        khoa = "None"
        nganh = "None"
        lop = "None"
        ten = "None"

    danh_sach = [truong, khoa, nganh, lop, ten]
    print("Nhập thông tin xong!")
    return danh_sach

def Cham_diem(bai_thi, dap_an, thong_tin):
    if bai_thi is not None and dap_an is not None and thong_tin is not None:
        print("Bắt đầu chấm điểm...")
        made = bai_thi[0]
        sbd = bai_thi[1]
        cau_hoi_bai_thi = bai_thi[2]
        cau_hoi_dap_an = dap_an[1]
        DA_bai_thi = bai_thi[3]
        DA_dung = dap_an[2]
        muc_diem = dap_an[3]
        diem = 0
        so_cau_dung = 0

        # Sửa phần này để infor chứa trực tiếp các giá trị
        infor = [thong_tin[0],     # TRƯỜNG
                thong_tin[1],     # KHOA
                thong_tin[2],     # NGÀNH
                thong_tin[3],     # LỚP
                thong_tin[4],     # TÊN
                sbd,              # SỐ BÁO DANH
                dap_an[0][0],     # MÔN
                dap_an[0][1],     # MÃ ĐỀ
                dap_an[0][2]]     # SỐ LƯỢNG CÂU HỎI
        
        if True:
            print("---")
            print(f"Thông tin cho biết từ SBD: {sbd}")
            print(f'Tên trường: {thong_tin[0]}')
            print(f'Tên khoa: {thong_tin[1]}')
            print(f'Tên ngành: {thong_tin[2]}')
            print(f'Tên lớp: {thong_tin[3]}')
            print(f'Họ tên: {thong_tin[4]}')
            print("---")
            print("Dựa mã đề và số báo danh , người này đã:")
            print(f"Thi môn: {dap_an[0][0]}")
            print(f"Làm mã đề: {dap_an[0][1]}")
            print(f"Tổng số câu hỏi: {dap_an[0][2]}")
            print("Đã làm các câu hỏi với đáp án như sau:")
        for i in range(dap_an[0][2]):
            #print(f"{bai_thi[2][i]}\t{bai_thi[3][i]}\t{dap_an[2][i]}\t{dap_an[3][i]}")
            if cau_hoi_bai_thi[i] == cau_hoi_dap_an[i]:
                print(f"{cau_hoi_bai_thi[i]}\t", end="")
                if DA_bai_thi[i] is not None:
                    print(f"{DA_bai_thi[i]}\t", end="")
                    print(f"{DA_dung[i]}\t", end="")
                    if DA_bai_thi[i] == DA_dung[i]:
                        print(muc_diem[i])
                        so_cau_dung += 1
                        diem += int(muc_diem[i])
                    else:
                        #diem.append(0)
                        print("")
                else:
                    print(f"\t{DA_dung[i]}\t", end="")
                    print(f"{0}\tKhông có đáp án trong bài thi (!)")
            else:
                print("\t", end="")
            #print(f"{cau_hoi_bai_thi[i]}\t{cau_hoi_dap_an[i]}\t{DA_bai_thi[i]}\t{DA_dung[i]}\t{muc_diem[i]}")

        ket_qua = [infor, cau_hoi_dap_an, DA_bai_thi, DA_dung, muc_diem, so_cau_dung, diem]
        print("Đã chấm xong !")
        return ket_qua
    else:
        return
# Hàm tao_file_excel mới (đã gộp)
def tao_file_excel(so_bao_danh, ket_qua, noi_luu_file):
    ten_file = f"KET_QUA - {so_bao_danh}.xlsx"
    duong_dan_day_du = os.path.join(noi_luu_file, ten_file)

    try:
        wb_ket_qua = openpyxl.Workbook()
        ws_ket_qua = wb_ket_qua.active

        ten_truong_hoc = ket_qua[0][0]
        ten_khoa_nganh_hoc = ket_qua[0][1]
        ten_nganh_hoc = ket_qua[0][2]
        ten_lop_hoc = ket_qua[0][3]
        ho_ten = ket_qua[0][4]
        ten_mon_thi = ket_qua[0][6]
        ma_de = ket_qua[0][7]
        tong_so_cau = ket_qua[0][8]

        for col_letter in ['A', 'C']:
            ws_ket_qua.column_dimensions[col_letter].width = 15
        ws_ket_qua.column_dimensions['B'].width = 30
        ws_ket_qua.column_dimensions['D'].width = 20

        ws_ket_qua['A1'] = "TRƯỜNG"
        ws_ket_qua['A2'] = "KHOA"
        ws_ket_qua['A3'] = "NGÀNH"
        ws_ket_qua['A4'] = "LỚP"
        ws_ket_qua['A5'] = "HỌ TÊN"
        ws_ket_qua['A6'] = "SBD"
        ws_ket_qua['A7'] = "MÔN THI"
        ws_ket_qua['A8'] = "MÃ ĐỀ"
        ws_ket_qua['A9'] = "TỔNG SỐ CÂU"
        ws_ket_qua['B1'] = ten_truong_hoc
        ws_ket_qua['B2'] = ten_khoa_nganh_hoc
        ws_ket_qua['B3'] = ten_nganh_hoc
        ws_ket_qua['B4'] = ten_lop_hoc
        ws_ket_qua['B5'] = ho_ten
        ws_ket_qua['B6'] = f'="{so_bao_danh}"'
        ws_ket_qua['B7'] = ten_mon_thi
        ws_ket_qua['B8'] = f'="{ma_de}"'
        ws_ket_qua['B9'] = tong_so_cau

        ws_ket_qua['A11'] = "BÀI LÀM:"
        ws_ket_qua['B11'] = "ĐÁP ÁN CHỌN"
        ws_ket_qua['C11'] = "ĐÁP ÁN ĐÚNG"
        ws_ket_qua['D11'] = "ĐIỂM"
        ws_ket_qua['D1'] = "SỐ CÂU LÀM ĐÚNG"
        ws_ket_qua['D2'] = f'=COUNTIF(D12:D{tong_so_cau+11},">0")&"/"&B9'
        ws_ket_qua['D4'] = "TỔNG ĐIỂM"
        ws_ket_qua['D5'] = f"=SUM(D12:D{tong_so_cau+11})"
        ws_ket_qua['D7'] = "THỜI GIAN"
        ws_ket_qua['D8'] = datetime.now().strftime("%d/%m/%Y - %H:%M:%S")

        red_font = Font(color='CC0000')
        green_font = Font(color='00A250')
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        list_of_cell_objects_center_alignment = ['D1','D2', 'D4', 'D5', 'D7', 'D8']
        list_of_cell_objects_bold = ['D1','D4','D7','A11','B11','C11','D11']
        cell_range_1 = ws_ket_qua['A1':'A9']
        cell_range_left_alignment = ws_ket_qua['A1':'B9']
        cell_range_center_alignment = ws_ket_qua['A11':'D11']

        start_row = 12
        end_row = 11 + tong_so_cau

        for row_of_cells in cell_range_1:
            for cell in row_of_cells:
                cell.font = bold_font
        for row_of_cells in cell_range_center_alignment:
            for cell in row_of_cells:
                cell.alignment = center_alignment
        for row_of_cells in cell_range_left_alignment:
            for cell in row_of_cells:
                cell.alignment = left_alignment
        for cell_address in list_of_cell_objects_center_alignment:
            ws_ket_qua[cell_address].alignment = center_alignment
        for cell_address in list_of_cell_objects_bold:
            ws_ket_qua[cell_address].font = bold_font

        for row_num in range(start_row, end_row + 1):
            list_idx = row_num - start_row
            try:
                cau_hoi_so_DA = ket_qua[1][list_idx]
                dap_an_bai_thi = ket_qua[2][list_idx]
                dap_an_DA = ket_qua[3][list_idx]
                diem_DA = ket_qua[4][list_idx]

                ws_ket_qua.cell(row=row_num, column=1).value = f"CÂU {cau_hoi_so_DA}"
                ws_ket_qua.cell(row=row_num, column=2).value = dap_an_bai_thi
                ws_ket_qua.cell(row=row_num, column=3).value = dap_an_DA

                if dap_an_bai_thi == dap_an_DA:
                    ws_ket_qua.cell(row=row_num, column=2).font = green_font
                else:
                    ws_ket_qua.cell(row=row_num, column=2).font = red_font

                ws_ket_qua.cell(row=row_num, column=1).alignment = center_alignment
                ws_ket_qua.cell(row=row_num, column=2).alignment = center_alignment
                ws_ket_qua.cell(row=row_num, column=3).alignment = center_alignment
                ws_ket_qua.cell(row=row_num, column=4).alignment = center_alignment

                ws_ket_qua.cell(row=row_num, column=4).value = f'=IF(AND(COUNTA(A{row_num}:C{row_num})=3,NOT(ISBLANK(B{row_num}))),IF(B{row_num}=C{row_num},{diem_DA},0),"")'

            except IndexError:
                print(f"{Fore.RED}LỖI INDEX: Chỉ số {list_idx} vượt quá phạm vi trong vòng lặp tại dòng Excel {row_num}. Kiểm tra lại số lượng câu trong file đáp án và dữ liệu bài thi.{Style.RESET_ALL}")
                break

        wb_ket_qua.save(duong_dan_day_du)
        #print(f"{Fore.GREEN}Đã tạo và lưu file '{Fore.CYAN}{ten_file}{Fore.GREEN}' tại {Fore.CYAN}{duong_dan_day_du}{Style.RESET_ALL}")
        return duong_dan_day_du

    except Exception as e:
        print(f"{Fore.RED}Đã xảy ra lỗi khi tạo hoặc ghi file Excel: {e}{Style.RESET_ALL}")
        return None
# --- XONG PHẦN GỘP HÀM ---
#-----------------------------------------------

# Thực thi
if __name__ == "__main__":
    file_mo_hinh = r"C:\Users\Y Bao\Downloads\Thu_muc_vao\AI_nhom_1.pt"
    thu_muc_anh = r"C:\Users\Y Bao\Downloads\Thu_muc_vao\Thu_muc_anh"
    file_danh_sach = r"C:\Users\Y Bao\Downloads\Thu_muc_vao\DS_lop.xlsx"
    file_dap_an = r"C:\Users\Y Bao\Downloads\Thu_muc_vao\Dap_an.xlsx"
    output_dir = r"C:\Users\Y Bao\Downloads\Thu_muc_ket_qua_ra"
    bai_thi = cham_bai(file_mo_hinh, thu_muc_anh, file_danh_sach, file_dap_an, output_dir, kich_thuoc_cat=(630, 510))
