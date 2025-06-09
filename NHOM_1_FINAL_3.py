import os
import time
from ultralytics import YOLO
import pandas as pd
import cv2
import numpy as np
from typing import Tuple, List, Optional
from tqdm import tqdm
import openpyxl
from colorama import init, Fore, Style
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image, ImageDraw, ImageFont, ImageTk
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading

# Khởi tạo colorama
init(autoreset=True)

luu_anh_cat = True
luu_anh_goc = True
luu_anh_cham = True
luu_anh_contours = True
luu_anh_nhi_phan = False

def chen_thong_tin_len_anh(anh, thong_tin):
    """
    Vẽ chữ tiếng Việt lên ảnh OpenCV sử dụng PIL/Pillow
    Chỉnh sửa trực tiếp ảnh gốc (in-place)
    """
    anh_rgb = cv2.cvtColor(anh, cv2.COLOR_BGR2RGB)
    pil_image = Image.fromarray(anh_rgb)
    draw = ImageDraw.Draw(pil_image)
    
    chu1 = f"Lớp: {thong_tin[0]}"
    chu2 = f"Họ tên: {thong_tin[1]}"
    chu3 = f"Số báo danh: {thong_tin[2]}"
    chu4 = f"Mã đề: {thong_tin[3]}"
    chu5 = f"Điểm: {thong_tin[4]}"
    chu6 = f"{thong_tin[5]}"
    
    try:
        font = ImageFont.truetype("arial.ttf", 14)
    except IOError:
        font = ImageFont.load_default()
    
    mau_chu = (130, 0, 0)  # Đỏ đậm
    vi_tri = [(20, 5), (20, 25), (20, 45), (20, 65), (20, 85), (420, 5)]
    cac_chu = [chu1, chu2, chu3, chu4, chu5, chu6]
    
    for i, chu in enumerate(cac_chu):
        draw.text(vi_tri[i], chu, font=font, fill=mau_chu)
    
    anh_cv2 = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
    anh[:] = anh_cv2

def tao_hinh_trong():
    h, w = 510, 630
    mau_bgr = (214, 227, 251)
    img = np.full((h, w, 3), mau_bgr, dtype=np.uint8)
    Text_1 = "No data !"
    font = cv2.FONT_HERSHEY_SIMPLEX
    co_chu = 2
    do_day = 4
    mau_chu = (0, 0, 192)
    (text_w, text_h), _ = cv2.getTextSize(Text_1, font, co_chu, do_day)
    vi_tri = ((w - text_w) // 2, (h + text_h) // 2)
    cv2.putText(img, Text_1, vi_tri, font, co_chu, mau_chu, do_day, cv2.LINE_AA)
    return img

def hien_thi_thoi_gian():
    thoi_gian_hien_tai = datetime.now()
    dinh_dang = thoi_gian_hien_tai.strftime("%Y_%m_%d - %H : %M : %S")
    return dinh_dang

def ReSize(img, ChieuCao):
    if ChieuCao is None:
        target_height = 500
    else:
        target_height = ChieuCao
    
    if img is None or img.shape[0] == 0:
        return tao_hinh_trong()
    
    original_height, original_width = img.shape[:2]
    scale = target_height / original_height
    new_width = int(original_width * scale)
    new_size = (new_width, target_height)
    interpolation_method = cv2.INTER_AREA if scale < 1 else cv2.INTER_LINEAR
    resized_img = cv2.resize(img, new_size, interpolation=interpolation_method)
    return resized_img

def tao_so_luong_cau_hoi(so_luong):
    if so_luong > 0:
        cau_hoi = []
        so_luong = round(so_luong)
        for i in range(1, so_luong + 1):
            cau_hoi.append(i)
        return cau_hoi

def to_mau_dap_an(anh_cham, ma_tran_dap_an, ket_qua_dap_an, dap_an_chon, dap_an_dung, to_thong_tin):
    for i in to_thong_tin:
        cv2.circle(anh_cham, i, 10, (98,250,202), 2)
    
    for hang_idx, hang in enumerate(ma_tran_dap_an):
        if hang_idx < len(dap_an_chon) and hang_idx < len(dap_an_dung):
            dap_an_hoc_sinh = dap_an_chon[hang_idx]
            dap_an_chuan = dap_an_dung[hang_idx]
            
            for cot_idx, (x, y) in enumerate(hang):
                if (hang_idx < len(ket_qua_dap_an) and 
                    cot_idx < len(ket_qua_dap_an[hang_idx]) and 
                    ket_qua_dap_an[hang_idx][cot_idx]):
                    
                    if dap_an_hoc_sinh == dap_an_chuan:
                        mau = (0, 255, 0)
                    else:
                        mau = (0, 0, 255)
                    
                    cv2.circle(anh_cham, (x, y), 10, mau, 2)

def cat_anh_co_dinh(img, toa_do_4_diem, kich_thuoc_out):
    if toa_do_4_diem is not None and len(toa_do_4_diem) == 4:
        points = np.array(toa_do_4_diem, dtype=np.float32)
        s = points.sum(axis=1)
        diff = np.diff(points, axis=1)
        tl = points[np.argmin(s)]
        br = points[np.argmax(s)]
        tr = points[np.argmin(diff)]
        bl = points[np.argmax(diff)]
        pts1 = np.float32([tl, tr, br, bl])
        width, height = kich_thuoc_out
        pts2 = np.float32([[0, 0], [width - 1, 0], [width - 1, height - 1], [0, height - 1]])
        matrix = cv2.getPerspectiveTransform(pts1, pts2)
        warped = cv2.warpPerspective(img, matrix, (width, height))
        print(f"Đã nắn phẳng phối cảnh và cắt ảnh với kích thước: {kich_thuoc_out}")
        return warped
    else:
        print("Không đủ 4 điểm để nắn phẳng phối cảnh.")
        return None

def nhan_dien_trac_nghiem(ten_file, anh_da_cat: np.ndarray) -> tuple:
    try:
        anh = anh_da_cat
        anh_xam = cv2.cvtColor(anh, cv2.COLOR_BGR2GRAY)
        anh_nhi_phan = cv2.adaptiveThreshold(anh_xam, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 11, 3)

        contours, _ = cv2.findContours(anh_nhi_phan, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        anh_contours = anh.copy()
        cv2.drawContours(anh_contours, contours, -1, (0, 255, 255), 2)
        cac_bubble = [
            (int(x), int(y))
            for contour in contours
            for (x, y), ban_kinh in [cv2.minEnclosingCircle(contour)]
            if 8 < ban_kinh < 25 and 0.5 < cv2.contourArea(contour) / (np.pi * ban_kinh**2 + 1e-5) < 1.5
        ]

        if not cac_bubble:
            print(f"{Fore.RED}Lỗi: Không tìm thấy bubble nào trong ảnh đã cắt")
            # Trả về các giá trị rỗng để tránh lỗi ở hàm gọi
            return anh.copy(), anh.copy(), anh_contours, "", "", [], [], [], [[] for _ in range(10)], [[] for _ in range(10)]

        toa_do = np.array(cac_bubble)
        x_gia_tri = toa_do[:, 0]
        x_min, x_max = x_gia_tri.min(), x_gia_tri.max()
        nguong_sbd = x_min + (x_max - x_min) * 0.3
        nguong_ma_de = x_min + (x_max - x_min) * 0.5

        nhom_sbd = [(x, y) for x, y in cac_bubble if x < nguong_sbd]
        nhom_ma_de = [(x, y) for x, y in cac_bubble if nguong_sbd <= x < nguong_ma_de]
        nhom_dap_an = [(x, y) for x, y in cac_bubble if x >= nguong_ma_de]
        
        # In cảnh báo theo code gốc
        if len(nhom_sbd) != 40:
            print(f"{Fore.RED}(!) CẢNH BÁO: các ô ở vùng SỐ BÁO DANH không được phát hiện đầy đủ, ảnh hưởng đến độ chính xác !")
            print(f"{Fore.YELLOW}--> Hãy thử chụp lại {Fore.BLUE}{ten_file}!{Fore.RESET}")
        if len(nhom_ma_de) != 20:
            print(f"{Fore.RED}(!) CẢNH BÁO: các ô ở vùng MÃ ĐỀ không được phát hiện đầy đủ, ảnh hưởng đến độ chính xác !")
            print(f"{Fore.YELLOW}--> Hãy thử chụp lại {Fore.BLUE}{ten_file}!{Fore.RESET}")
        if len(nhom_dap_an) != 40:
            print(f"{Fore.RED}(!) CẢNH BÁO: các ô ở vùng ĐÁP ÁN không được phát hiện đầy đủ, ảnh hưởng đến độ chính xác !")
            print(f"{Fore.YELLOW}--> Hãy thử chụp lại {Fore.BLUE}{ten_file}!{Fore.RESET}")


        def sap_xep_bubble(bubbles: List[Tuple[int, int]], so_hang: int = 10, so_cot: int = 4) -> List[List[Tuple[int, int]]]:
            if not bubbles: return [[] for _ in range(so_hang)]
            bubbles = sorted(bubbles, key=lambda b: b[1])
            khoang_cach_y = (bubbles[-1][1] - bubbles[0][1]) / (so_hang - 1) if so_hang > 1 else 0
            hang = [[] for _ in range(so_hang)]
            for x, y in bubbles:
                chi_so_hang = min(so_hang - 1, max(0, int(round((y - bubbles[0][1]) / khoang_cach_y)) if khoang_cach_y > 0 else 0))
                hang[chi_so_hang].append((x, y))
            for i in range(so_hang):
                hang[i].sort(key=lambda b: b[0])
            return hang

        ma_tran_sbd = sap_xep_bubble(nhom_sbd, 10, 4)
        ma_tran_ma_de = sap_xep_bubble(nhom_ma_de, 10, 2)
        ma_tran_dap_an = sap_xep_bubble(nhom_dap_an, 10, 4)
        to_thong_tin = []
        to_dap_an = []

        def doc_to(anh_nhi_phan: np.ndarray, ma_tran: List[List[Tuple[int, int]]], ban_kinh: int = 10, nguong: float = 0.5) -> List[List[bool]]:
            ket_qua = []
            for hang in ma_tran:
                dong = []
                for x, y in hang:
                    vung = anh_nhi_phan[y - ban_kinh:y + ban_kinh, x - ban_kinh:x + ban_kinh]
                    ty_le = cv2.countNonZero(vung) / (vung.size + 1e-5) if vung.size else 0
                    dong.append(ty_le > nguong)
                ket_qua.append(dong)
            return ket_qua

        ket_qua_sbd = doc_to(anh_nhi_phan, ma_tran_sbd)
        ket_qua_ma_de = doc_to(anh_nhi_phan, ma_tran_ma_de)
        ket_qua_dap_an = doc_to(anh_nhi_phan, ma_tran_dap_an)

        def doc_so_bao_danh(ma_tran_kq: List[List[bool]]) -> str:
            sbd = ""
            num_cols = len(ma_tran_kq[0]) if ma_tran_kq and ma_tran_kq[0] else 0
            for cot in range(min(4, num_cols)):
                gia_tri = "_"
                for hang_idx, hang in enumerate(ma_tran_kq):
                    if cot < len(hang) and hang[cot]:
                        gia_tri = str(hang_idx)
                        break
                sbd += gia_tri
            return sbd

        def doc_ma_de(ma_tran_kq: List[List[bool]]) -> str:
            ma_de = ""
            num_cols = len(ma_tran_kq[0]) if ma_tran_kq and ma_tran_kq[0] else 0
            for cot in range(min(2, num_cols)):
                gia_tri = "_"
                for hang_idx, hang in enumerate(ma_tran_kq):
                     if cot < len(hang) and hang[cot]:
                        gia_tri = str(hang_idx)
                        break
                ma_de += gia_tri
            return ma_de

        def doc_dap_an(ma_tran_kq: List[List[bool]]) -> List[str]:
            dap_an = []
            for hang in ma_tran_kq:
                lua_chon = ""
                if len(hang) == 4:
                    if hang[0]: lua_chon += "A"
                    if hang[1]: lua_chon += "B"
                    if hang[2]: lua_chon += "C"
                    if hang[3]: lua_chon += "D"
                dap_an.append(lua_chon or "_")
            return dap_an

        so_bao_danh = doc_so_bao_danh(ket_qua_sbd)
        ma_de = doc_ma_de(ket_qua_ma_de)
        dap_an = doc_dap_an(ket_qua_dap_an)
        
        anh_debug = anh.copy()
        def ve_bubble(anh, ma_tran, ket_qua, mau_to=(0, 0, 255), mau_khong_to=(69, 255, 255)):
            for i, hang in enumerate(ma_tran):
                for j, (x, y) in enumerate(hang):
                    if i < len(ket_qua) and j < len(ket_qua[i]):
                        mau = mau_to if ket_qua[i][j] else mau_khong_to
                        cv2.circle(anh, (x, y), 10, mau, 2)
        
        ve_bubble(anh_debug, ma_tran_sbd, ket_qua_sbd)
        ve_bubble(anh_debug, ma_tran_ma_de, ket_qua_ma_de)
        ve_bubble(anh_debug, ma_tran_dap_an, ket_qua_dap_an)

        def lay_danh_sach_to(ma_tran, ket_qua):
            danh_sach = []
            for i, hang in enumerate(ma_tran):
                for j, (x,y) in enumerate(hang):
                     if i < len(ket_qua) and j < len(ket_qua[i]) and ket_qua[i][j]:
                        danh_sach.append((x,y))
            return danh_sach

        to_thong_tin.extend(lay_danh_sach_to(ma_tran_sbd, ket_qua_sbd))
        to_thong_tin.extend(lay_danh_sach_to(ma_tran_ma_de, ket_qua_ma_de))
        to_dap_an.extend(lay_danh_sach_to(ma_tran_dap_an, ket_qua_dap_an))
        
        return anh_nhi_phan, anh_debug, anh_contours, so_bao_danh, ma_de, dap_an, to_thong_tin, to_dap_an, ma_tran_dap_an, ket_qua_dap_an

    except Exception as e:
        print(f"Lỗi không xác định trong nhan_dien_trac_nghiem: {e}")
        anh_trong = tao_hinh_trong()
        return anh_trong, anh_trong, anh_trong, "", "", [], [], [], [[]], [[]]

def Nhap_file_dap_an(file_DA):
    df_da = pd.read_excel(file_DA, header=None)
    mon_thi = df_da.iloc[df_da.iloc[:, 0][(df_da.iloc[:, 0] == "Môn thi")].index,1].item()
    ma_de_dap_an = df_da.iloc[df_da.iloc[:, 0][(df_da.iloc[:, 0] == "Mã đề")].index,1].item()
    tong_so_cau = df_da.iloc[df_da.iloc[:, 0][(df_da.iloc[:, 0] == "Tổng số câu")].index,1].item()
    dap_an = [[mon_thi, ma_de_dap_an, tong_so_cau], [], [], []]
    
    start_index = df_da.iloc[:, 0][(df_da.iloc[:, 0] == "Câu hỏi")].index[0] + 1
    for i in range(tong_so_cau):
        row_index = start_index + i
        cau_hoi_bai_thi_value = df_da.iloc[row_index,0]
        dap_an_dung_value = df_da.iloc[row_index,1]
        muc_diem_value = df_da.iloc[row_index,2]      
        dap_an[1].append(cau_hoi_bai_thi_value)
        dap_an[2].append(dap_an_dung_value)
        dap_an[3].append(muc_diem_value)
    print("Nhập file đáp án xong !")
    return dap_an

def Tim_thong_tin(file_danh_sach, SBD_can_tim):
    df_ds = pd.read_excel(file_danh_sach, header=None)
    last_valid = df_ds[1].last_valid_index()
    if last_valid is None:
        last_valid = 0
    ds_sbd_series = df_ds.loc[7:last_valid, 1]
    
    # Chuyển đổi SBD sang chuỗi để so sánh
    ds_sbd_values = ds_sbd_series.astype(str).values

    if "_" not in SBD_can_tim and SBD_can_tim in ds_sbd_values:
        truong = df_ds.loc[df_ds[0] == "TRƯỜNG", 1].values[0]
        khoa = df_ds.loc[df_ds[0] == "KHOA", 1].values[0]
        nganh = df_ds.loc[df_ds[0] == "NGÀNH", 1].values[0]
        lop = df_ds.loc[df_ds[0] == "LỚP", 1].values[0]
        
        dong_ten = df_ds[df_ds[1].astype(str) == SBD_can_tim].index[0]
        ten = df_ds.loc[dong_ten, 0]
    else:
        truong = khoa = nganh = lop = ten = "Không tìm thấy"

    danh_sach = [truong, khoa, nganh, lop, ten]
    print("Nhập thông tin xong!")
    return danh_sach

def Cham_diem(bai_thi, dap_an, thong_tin):
    if bai_thi is None or dap_an is None or thong_tin is None:
        print(f"{Fore.RED}Lỗi: Thiếu thông tin đầu vào cho hàm chấm điểm.")
        return None
    print("Bắt đầu chấm điểm...")
    made, sbd, cau_hoi_bai_thi, DA_bai_thi = bai_thi
    tong_so_cau_dap_an, cau_hoi_dap_an, DA_dung, muc_diem = dap_an[0][2], dap_an[1], dap_an[2], dap_an[3]
    diem, so_cau_dung = 0, 0

    infor = thong_tin + [sbd, dap_an[0][0], dap_an[0][1], dap_an[0][2]]
    
    print("---")
    print(f"Thông tin SV: {thong_tin[4]} - SBD: {sbd} - Lớp: {thong_tin[3]}")
    print(f"Thi môn: {dap_an[0][0]} - Làm mã đề: {made}")
    print(f"Tổng số câu hỏi: {tong_so_cau_dap_an}")
    print("---")

    for i in range(tong_so_cau_dap_an):
        if i < len(DA_bai_thi) and i < len(DA_dung):
            print(f"Câu {cau_hoi_dap_an[i]}: Chọn '{DA_bai_thi[i]}', Đáp án đúng '{DA_dung[i]}'", end=" -> ")
            if DA_bai_thi[i] == DA_dung[i]:
                so_cau_dung += 1
                current_score = int(muc_diem[i])
                diem += current_score
                print(f"Đúng (+{current_score}đ)")
            else:
                print("Sai")
    
    ket_qua = [infor, cau_hoi_dap_an, DA_bai_thi, DA_dung, muc_diem, so_cau_dung, diem]
    print("Đã chấm xong !")
    return ket_qua

def tao_file_excel(so_bao_danh, ket_qua, noi_luu_file):
    if not ket_qua:
        print(f"{Fore.RED}Không có kết quả để tạo file Excel.{Style.RESET_ALL}")
        return None
        
    ten_file = f"KET_QUA - {so_bao_danh}.xlsx"
    duong_dan_day_du = os.path.join(noi_luu_file, ten_file)

    try:
        wb_ket_qua = openpyxl.Workbook()
        ws_ket_qua = wb_ket_qua.active

        ten_truong_hoc = ket_qua[0][0]
        ten_khoa_nganh_hoc = ket_qua[0][1]
        ten_nganh_hoc = ket_qua[0][2]
        ten_lop_hoc = ket_qua[0][3]
        ho_ten = ket_qua[0][4]
        sbd_excel = ket_qua[0][5]
        ten_mon_thi = ket_qua[0][6]
        ma_de = ket_qua[0][7]
        tong_so_cau = ket_qua[0][8]

        for col_letter in ['A', 'C']: ws_ket_qua.column_dimensions[col_letter].width = 15
        ws_ket_qua.column_dimensions['B'].width = 30
        ws_ket_qua.column_dimensions['D'].width = 20

        ws_ket_qua['A1'] = "TRƯỜNG"; ws_ket_qua['B1'] = ten_truong_hoc
        ws_ket_qua['A2'] = "KHOA"; ws_ket_qua['B2'] = ten_khoa_nganh_hoc
        ws_ket_qua['A3'] = "NGÀNH"; ws_ket_qua['B3'] = ten_nganh_hoc
        ws_ket_qua['A4'] = "LỚP"; ws_ket_qua['B4'] = ten_lop_hoc
        ws_ket_qua['A5'] = "HỌ TÊN"; ws_ket_qua['B5'] = ho_ten
        ws_ket_qua['A6'] = "SBD"; ws_ket_qua['B6'] = f'="{sbd_excel}"'
        ws_ket_qua['A7'] = "MÔN THI"; ws_ket_qua['B7'] = ten_mon_thi
        ws_ket_qua['A8'] = "MÃ ĐỀ"; ws_ket_qua['B8'] = f'="{ma_de}"'
        ws_ket_qua['A9'] = "TỔNG SỐ CÂU"; ws_ket_qua['B9'] = tong_so_cau

        ws_ket_qua['A11'] = "BÀI LÀM:"; ws_ket_qua['B11'] = "ĐÁP ÁN CHỌN"
        ws_ket_qua['C11'] = "ĐÁP ÁN ĐÚNG"; ws_ket_qua['D11'] = "ĐIỂM"
        ws_ket_qua['D1'] = "SỐ CÂU LÀM ĐÚNG"; ws_ket_qua['D2'] = f'=COUNTIF(D12:D{tong_so_cau+11},">0")&"/"&B9'
        ws_ket_qua['D4'] = "TỔNG ĐIỂM"; ws_ket_qua['D5'] = f"=SUM(D12:D{tong_so_cau+11})"
        ws_ket_qua['D7'] = "THỜI GIAN"; ws_ket_qua['D8'] = datetime.now().strftime("%d/%m/%Y - %H:%M:%S")

        red_font, green_font, bold_font = Font(color='CC0000'), Font(color='00A250'), Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        bold_cells = ['D1','D4','D7','A11','B11','C11','D11'] + [f'A{i}' for i in range(1, 10)]
        for cell_ref in bold_cells: ws_ket_qua[cell_ref].font = bold_font
        for row in ws_ket_qua['A11':'D11']:
            for cell in row: cell.alignment = center_align
        for row in ws_ket_qua['A1':'B9']:
            for cell in row: cell.alignment = left_align
        for cell_ref in ['D1','D2', 'D4', 'D5', 'D7', 'D8']: ws_ket_qua[cell_ref].alignment = center_align
        
        start_row, end_row = 12, 11 + tong_so_cau
        for row_num in range(start_row, end_row + 1):
            idx = row_num - start_row
            try:
                if idx < len(ket_qua[1]) and idx < len(ket_qua[2]) and idx < len(ket_qua[3]) and idx < len(ket_qua[4]):
                    ws_ket_qua.cell(row=row_num, column=1, value=f"CÂU {ket_qua[1][idx]}")
                    ws_ket_qua.cell(row=row_num, column=2, value=ket_qua[2][idx])
                    ws_ket_qua.cell(row=row_num, column=3, value=ket_qua[3][idx])
                    
                    if ket_qua[2][idx] == ket_qua[3][idx]: ws_ket_qua.cell(row=row_num, column=2).font = green_font
                    else: ws_ket_qua.cell(row=row_num, column=2).font = red_font
                    
                    ws_ket_qua.cell(row=row_num, column=4, value=f'=IF(B{row_num}=C{row_num},{ket_qua[4][idx]},0)')
                    for col in range(1, 5): ws_ket_qua.cell(row=row_num, column=col).alignment = center_align
            except IndexError:
                print(f"{Fore.RED}LỖI INDEX: Chỉ số {idx} vượt quá phạm vi. Dừng ghi file Excel.{Style.RESET_ALL}")
                break

        wb_ket_qua.save(duong_dan_day_du)
        return duong_dan_day_du

    except Exception as e:
        print(f"{Fore.RED}Đã xảy ra lỗi khi tạo file Excel: {e}{Style.RESET_ALL}")
        return None

# === HÀM XỬ LÝ CHÍNH ĐƯỢC TÍCH HỢP VÀO GUI ===
def process_single_image(image_path, model_path, class_list_path, answer_key_path, output_dir, gui_app, kich_thuoc_cat=(630, 510)):
    # Khởi tạo các biến ảnh trả về
    anh_danh_dau_goc = None
    anh_da_cat = None
    anh_nhi_phan = None
    anh_contours = None
    anh_debug = None
    anh_cham = None

    # Tạo các ảnh trống để hiển thị nếu có lỗi
    img_placeholder = tao_hinh_trong()
    try:
        
        # 1. Tạo thư mục kết quả (nếu chưa có)
        anh_goc_dir = os.path.join(output_dir, 'anh_goc_danh_dau')
        anh_cat_dir = os.path.join(output_dir, 'anh_da_cat_va_nhan_dien')
        output_anh_nhi_phan_dir = os.path.join(output_dir, 'anh_nhi_phan')
        anh_cham_dir = os.path.join(output_dir, 'anh_da_cham')
        excel_dir = os.path.join(output_dir, 'file_excel')
        for d in [anh_goc_dir, anh_cat_dir, output_anh_nhi_phan_dir, anh_cham_dir, excel_dir]:
            os.makedirs(d, exist_ok=True)
            
        base_name = os.path.basename(image_path)
        print(f"\n{Fore.YELLOW}" + "-" * 50 + f"{Fore.RESET}")

        # 2. Thực hiện dự đoán
        model = YOLO(model_path)
        results = model.predict(image_path, save=False, verbose=False)
        result = results[0]

        boxes = result.boxes.xyxy
        toa_do_all = []
        print(f"Đã xử lý ảnh: '{Fore.BLUE}{base_name}{Fore.RESET}' - Đã phát hiện {Fore.YELLOW}{len(boxes)}{Fore.RESET} đối tượng:")
        gui_app.log_terminal(f"Đang xử lý '{base_name}': {len(boxes)} đối tượng.")
        
        for box in boxes:
            xmin, ymin, xmax, ymax = [int(coord) for coord in box.tolist()]
            center_x = round((xmin + xmax) / 2)
            center_y = round((ymin + ymax) / 2)
            toa_do_all.append((center_x, center_y))
            print(f"  - Tâm: ({center_x}, {center_y})")

        # Đọc ảnh gốc và vẽ
        anh_goc = cv2.imread(image_path)
        anh_danh_dau_goc = result.plot()
        gui_app.update_original_image(anh_danh_dau_goc)

        # Lưu ảnh gốc đã đánh dấu (theo logic gốc)
        if luu_anh_goc:
            output_goc_path = os.path.join(anh_goc_dir, f'ANH_GOC - {base_name}')
            cv2.imwrite(output_goc_path, anh_danh_dau_goc)
            print(f"{Fore.GREEN}--> Ảnh gốc đã đánh dấu được lưu tại: {Fore.CYAN}{output_goc_path}")


        # 3. Cắt ảnh
        bo_qua_anh = False
        if len(toa_do_all) == 4:
            anh_da_cat = cat_anh_co_dinh(anh_goc.copy(), toa_do_all, kich_thuoc_cat)
        else:
            print(f"{Fore.RED}(!) CẢNH BÁO: Phát hiện {len(toa_do_all)}/4 đối tượng. Bỏ qua xử lý ảnh này.")
            gui_app.log_terminal(f"(!) CẢNH BÁO: Phát hiện {len(toa_do_all)}/4 góc. Bỏ qua ảnh.")
            bo_qua_anh = True

        # 4. Nếu không bỏ qua, tiếp tục xử lý
        if not bo_qua_anh and anh_da_cat is not None:
            (anh_nhi_phan, anh_debug, anh_contours, sbd, ma_de, dap_an_chon, 
             to_thong_tin, to_dap_an, ma_tran_dap_an, ket_qua_dap_an) = nhan_dien_trac_nghiem(base_name, anh_da_cat)

            print(f"Kết quả nhận diện: Số báo danh: {sbd} | Mã đề: {ma_de}")
            gui_app.log_terminal(f"Nhận diện: SBD={sbd}, Mã đề={ma_de}")
            
            # 5. Chấm điểm (theo logic gốc)
            dap_an = Nhap_file_dap_an(answer_key_path)
            thong_tin = Tim_thong_tin(class_list_path, sbd)
            bai_thi = [ma_de, sbd, tao_so_luong_cau_hoi(dap_an[0][2]), dap_an_chon]
            ket_qua = Cham_diem(bai_thi, dap_an, thong_tin)
            
            if ket_qua:
                print(f"Kết quả chấm điểm: Số câu đúng: {ket_qua[5]}, Tổng điểm: {ket_qua[6]}")
                gui_app.log_terminal(f"Kết quả: {ket_qua[5]} câu đúng, Điểm: {ket_qua[6]}")

                # 6. Tạo ảnh đã chấm
                anh_cham = anh_da_cat.copy()
                dap_an_dung_list = dap_an[2]
                to_mau_dap_an(anh_cham, ma_tran_dap_an, ket_qua_dap_an, dap_an_chon, dap_an_dung_list, to_thong_tin)
                
                info_text = (
                    ket_qua[0][3], # Lớp
                    ket_qua[0][4], # Họ tên
                    sbd,
                    ma_de,
                    ket_qua[6],    # Điểm
                    hien_thi_thoi_gian()
                )
                chen_thong_tin_len_anh(anh_cham, info_text)
                
                # 7. Lưu file Excel (theo logic gốc)
                excel_path = tao_file_excel(sbd, ket_qua, excel_dir)
                if excel_path:
                    print(f"{Fore.GREEN}--> Đã lưu file EXCEL tại: {Fore.CYAN}{excel_path}")
                
                # 8. Lưu các ảnh trung gian (theo logic gốc)
                if luu_anh_cat and anh_debug is not None:
                    path = os.path.join(anh_cat_dir, f'ANH_CAT - {sbd} - {base_name}')
                    cv2.imwrite(path, anh_debug)
                    print(f"{Fore.GREEN}--> Đã lưu ảnh nhận diện tại: {Fore.CYAN}{path}")
                if luu_anh_nhi_phan and anh_nhi_phan is not None:
                    path = os.path.join(output_anh_nhi_phan_dir, f'NHI_PHAN - {sbd} - {base_name}')
                    cv2.imwrite(path, anh_nhi_phan)
                    print(f"{Fore.GREEN}--> Đã lưu ảnh nhị phân tại: {Fore.CYAN}{path}")
                if luu_anh_cham and anh_cham is not None:
                    path = os.path.join(anh_cham_dir, f'ANH_CHAM - {sbd} - {base_name}')
                    cv2.imwrite(path, anh_cham)
                    print(f"{Fore.GREEN}--> Đã lưu ảnh đã chấm tại: {Fore.CYAN}{path}")
            else:
                gui_app.log_terminal("Lỗi trong quá trình chấm điểm.")
        else:
            # Nếu bỏ qua, các ảnh xử lý sẽ là ảnh trống
            anh_da_cat = img_placeholder
            anh_nhi_phan = img_placeholder
            anh_contours = img_placeholder
            anh_debug = img_placeholder
            anh_cham = img_placeholder

    except Exception as e:
        print(f"{Fore.RED}Lỗi nghiêm trọng trong hàm xử lý ảnh: {e}")
        gui_app.log_terminal(f"Lỗi: {e}")
        anh_da_cat = img_placeholder; anh_nhi_phan = img_placeholder
        anh_nhi_phan = img_placeholder; anh_debug = img_placeholder; anh_cham = img_placeholder
    
    finally:
        # ---- CẬP NHẬT GIAO DIỆN ----
        gui_app.update_processed_image(0, anh_da_cat)
        gui_app.update_processed_image(1, anh_nhi_phan)
        gui_app.update_processed_image(2, anh_debug)
        gui_app.update_processed_image(3, anh_cham)
        print(f"\n{Fore.YELLOW}=== ĐÃ XỬ LÝ XONG ẢNH {base_name} ==={Style.RESET_ALL}\n")

# === PHẦN GIAO DIỆN ===
class ImageProcessingGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Hệ thống chấm trắc nghiệm tự động - NHÓM 1 - CNTT K22")
        self.master.geometry("1300x800")
        self.master.minsize(1200, 750)
        self.master.configure(bg='#f0f0f0')

        self.auto_active = False
        self.running = False

        self.on_import_model = None
        self.on_import_folder = None
        self.on_import_list = None
        self.on_import_answer = None
        self.on_run_clicked = None
        self.on_auto_clicked = None
        self.on_auto_stopped = None
        
        self.create_widgets()

    def create_widgets(self):
        main_frame = tk.Frame(self.master, bg='#f0f0f0')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        left_frame = tk.Frame(main_frame, bg='white', relief=tk.RAISED, bd=2, width=220)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left_frame.pack_propagate(False)

        title_label = tk.Label(left_frame, text="BẢNG ĐIỀU KHIỂN", font=('Arial', 14, 'bold'), bg='white', fg='#003366')
        title_label.pack(pady=20, padx=10)

        menu_buttons_info = [
            ("Nhập file mô hình", self.import_model_callback),
            ("Nhập thư mục ảnh", self.import_folder_callback),
            ("Nhập file danh sách", self.import_list_callback),
            ("Nhập file đáp án", self.import_answer_callback)
        ]

        for text, command in menu_buttons_info:
            btn = tk.Button(left_frame, text=text, height=2, font=('Arial', 10), 
                           bg='#e1e1e1', relief=tk.RAISED, command=command)
            btn.pack(pady=5, padx=10, fill=tk.X)
        
        right_frame = tk.Frame(main_frame, bg='#f0f0f0')
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        top_frame = tk.Frame(right_frame, bg='#f0f0f0')
        top_frame.pack(fill=tk.X, pady=(0, 10))
        
        terminal_container = tk.LabelFrame(top_frame, text="Nhật ký hoạt động", font=('Arial', 10, 'bold'), bg='white', padx=5, pady=5)
        terminal_container.pack(fill=tk.BOTH, expand=True, side=tk.LEFT, padx=(0, 10))

        terminal_frame = tk.Frame(terminal_container, bg='black')
        terminal_frame.pack(fill=tk.BOTH, expand=True)
        self.terminal_text = tk.Text(terminal_frame, height=10, font=('Consolas', 9), bg="#F0F0F0", fg="#020202", wrap=tk.WORD, insertbackground='white')
        scrollbar = tk.Scrollbar(terminal_frame, orient=tk.VERTICAL, command=self.terminal_text.yview)
        self.terminal_text.configure(yscrollcommand=scrollbar.set)
        self.terminal_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        control_frame = tk.Frame(top_frame, bg='#f0f0f0')
        control_frame.pack(side=tk.RIGHT, fill=tk.Y)

        self.auto_btn = tk.Button(control_frame, text="TỰ ĐỘNG", width=12, height=2, font=('Arial', 10, 'bold'), bg='#2196F3', fg='white', command=self.auto_clicked)
        self.auto_btn.pack(side=tk.TOP, padx=5, pady=(0,5), fill=tk.BOTH, expand=True)
        self.run_btn = tk.Button(control_frame, text="CHẠY", width=12, height=2, font=('Arial', 10, 'bold'), bg='#4CAF50', fg='white', command=self.run_clicked)
        self.run_btn.pack(side=tk.BOTTOM, padx=5, pady=(5,0), fill=tk.BOTH, expand=True)

        bottom_frame = tk.Frame(right_frame, bg='#f0f0f0')
        bottom_frame.pack(fill=tk.BOTH, expand=True)
        self.create_image_grid(bottom_frame)

    def create_image_grid(self, parent):
        left_img_frame = tk.LabelFrame(parent, text="ẢNH GỐC (ĐÃ NHẬN DIỆN)", font=('Arial', 12, 'bold'), bg='white', relief=tk.RAISED, bd=2)
        left_img_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        self.original_img_label = tk.Label(left_img_frame, bg='#f8f8f8', relief=tk.SUNKEN, bd=2)
        self.original_img_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        right_img_frame = tk.Frame(parent, bg='#f0f0f0')
        right_img_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.processed_img_labels = {}
        positions = [(0, 0), (0, 1), (1, 0), (1, 1)]
        titles = ["Ảnh đã cắt", "Ảnh nhị phân", "Ảnh nhận diện", "Ảnh đã chấm"]

        for i, (row, col) in enumerate(positions):
            frame = tk.LabelFrame(right_img_frame, text=titles[i], font=('Arial', 10, 'bold'), bg='white', relief=tk.RAISED, bd=2)
            frame.grid(row=row, column=col, sticky='nsew', padx=2, pady=2)
            img_label = tk.Label(frame, bg='#f8f8f8', relief=tk.SUNKEN, bd=2)
            img_label.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            self.processed_img_labels[i] = img_label

        for i in range(2):
            right_img_frame.grid_rowconfigure(i, weight=1)
            right_img_frame.grid_columnconfigure(i, weight=1)

    def import_model_callback(self):
        if self.on_import_model: self.on_import_model()
    def import_folder_callback(self):
        if self.on_import_folder: self.on_import_folder()
    def import_list_callback(self):
        if self.on_import_list: self.on_import_list()
    def import_answer_callback(self):
        if self.on_import_answer: self.on_import_answer()
    
    def run_clicked(self):
        if self.running or self.auto_active: return
        self.running = True
        self.run_btn.configure(text="ĐANG CHẠY...", state='disabled', bg='#ff9800')
        if self.on_run_clicked: self.on_run_clicked()

    def _reset_run_button(self):
        self.running = False
        self.run_btn.configure(text="CHẠY", state='normal', bg='#4CAF50')

    def auto_clicked(self):
        if self.running: return
        if not self.auto_active:
            self.auto_active = True
            self.auto_btn.configure(text="DỪNG", bg='#f44336')
            self.run_btn.configure(state='disabled')
            self.log_terminal("Chế độ TỰ ĐỘNG đã được BẬT.")
            if self.on_auto_clicked: self.on_auto_clicked()
        else:
            self.auto_active = False
            self.auto_btn.configure(text="TỰ ĐỘNG", bg='#2196F3')
            self.run_btn.configure(state='normal')
            self.log_terminal("Chế độ TỰ ĐỘNG đã được TẮT.")
            if self.on_auto_stopped: self.on_auto_stopped()
    
    def stop_auto_mode(self):
        self.auto_active = False
        self.auto_btn.configure(text="TỰ ĐỘNG", bg='#2196F3')
        self.run_btn.configure(state='normal')

    def log_terminal(self, message):
        self.terminal_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.terminal_text.see(tk.END)
        self.master.update_idletasks()

    def update_image(self, label, cv_image):
        if cv_image is None: return
        try:
            self.master.update_idletasks()
            label_w, label_h = label.winfo_width(), label.winfo_height()
            if label_w <= 1 or label_h <= 1: return 

            img_h, img_w = cv_image.shape[:2]
            aspect_ratio = img_w / img_h
            
            new_w = label_w
            new_h = int(new_w / aspect_ratio)
            if new_h > label_h:
                new_h = label_h
                new_w = int(new_h * aspect_ratio)

            resized_img = cv2.resize(cv_image, (new_w, new_h), interpolation=cv2.INTER_AREA)
            
            img_rgb = cv2.cvtColor(resized_img, cv2.COLOR_BGR2RGB)
            photo = ImageTk.PhotoImage(image=Image.fromarray(img_rgb))
            
            label.configure(image=photo)
            label.image = photo
        except Exception as e:
            self.log_terminal(f"Lỗi hiển thị ảnh: {e}")

    def update_original_image(self, cv_image): self.update_image(self.original_img_label, cv_image)
    def update_processed_image(self, index, cv_image):
        if index in self.processed_img_labels: self.update_image(self.processed_img_labels[index], cv_image)
    def set_callbacks(self, callbacks):
        self.on_import_model = callbacks.get('on_import_model')
        self.on_import_folder = callbacks.get('on_import_folder')
        self.on_import_list = callbacks.get('on_import_list')
        self.on_import_answer = callbacks.get('on_import_answer')
        self.on_run_clicked = callbacks.get('on_run_clicked')
        self.on_auto_clicked = callbacks.get('on_auto_clicked')
        self.on_auto_stopped = callbacks.get('on_auto_stopped')

# === LỚP ĐIỀU KHIỂN CHÍNH (APP CONTROLLER) ===
class AppController:
    def __init__(self, master):
        self.master = master
        self.gui = ImageProcessingGUI(master)
        
        self.model_path = None
        self.image_folder = None
        self.class_list_path = None
        self.answer_key_path = None
        self.output_dir = 'ket_qua' # Thư mục kết quả mặc định
        
        self.image_files = []
        self.current_image_index = 0
        self.is_auto_running = False
        
        callbacks = {
            'on_import_model': self.select_model_file,
            'on_import_folder': self.select_image_folder,
            'on_import_list': self.select_class_list_file,
            'on_import_answer': self.select_answer_key_file,
            'on_run_clicked': self.run_single_step,
            'on_auto_clicked': self.start_auto_processing,
            'on_auto_stopped': self.stop_auto_processing
        }
        self.gui.set_callbacks(callbacks)
        #self.gui.log_terminal("Chào mừng! Vui lòng nhập các file cần thiết.")
        
    def _validate_inputs(self):
        """Kiểm tra xem tất cả các file cần thiết đã được nhập chưa"""
        if not self.model_path:
            messagebox.showerror("Thiếu thông tin", "Vui lòng nhập file mô hình (.pt)!"); return False
        if not self.image_folder:
            messagebox.showerror("Thiếu thông tin", "Vui lòng nhập thư mục ảnh!"); return False
        if not self.class_list_path:
            messagebox.showerror("Thiếu thông tin", "Vui lòng nhập file danh sách lớp (.xlsx)!"); return False
        if not self.answer_key_path:
            messagebox.showerror("Thiếu thông tin", "Vui lòng nhập file đáp án (.xlsx)!"); return False
        if not self.image_files:
            messagebox.showinfo("Thông báo", "Không tìm thấy ảnh nào trong thư mục đã chọn."); return False
        return True

    def select_model_file(self):
        path = filedialog.askopenfilename(title="Chọn file mô hình", filetypes=[("Model files", "*.pt")])
        if path:
            self.model_path = path
            self.gui.log_terminal(f"Đã chọn mô hình: {os.path.basename(path)}")
    
    def select_image_folder(self):
        path = filedialog.askdirectory(title="Chọn thư mục chứa ảnh")
        if path:
            self.image_folder = path
            self.scan_image_folder()
            # Gợi ý thư mục lưu kết quả
            parent_dir = os.path.dirname(path)
            self.output_dir = os.path.join(parent_dir, 'ket_qua_cham_bai')
            self.gui.log_terminal(f"Thư mục kết quả sẽ được lưu tại: {self.output_dir}")
    
    def select_class_list_file(self):
        path = filedialog.askopenfilename(title="Chọn file danh sách lớp", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.class_list_path = path
            self.gui.log_terminal(f"Đã chọn danh sách: {os.path.basename(path)}")
            
    def select_answer_key_file(self):
        path = filedialog.askopenfilename(title="Chọn file đáp án", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.answer_key_path = path
            self.gui.log_terminal(f"Đã chọn đáp án: {os.path.basename(path)}")

    def scan_image_folder(self):
        """Quét thư mục ảnh và cập nhật danh sách file"""
        self.image_files = []
        valid_extensions = ('.jpg', '.jpeg', '.png', '.bmp')
        if not self.image_folder: return
        for f in os.listdir(self.image_folder):
            if f.lower().endswith(valid_extensions):
                self.image_files.append(os.path.join(self.image_folder, f))
        
        self.current_image_index = 0
        if self.image_files:
            self.gui.log_terminal(f"Đã chọn thư mục ảnh: {self.image_folder}")
            self.gui.log_terminal(f"Tìm thấy {len(self.image_files)} ảnh. Sẵn sàng để chấm.")
        else:
            self.gui.log_terminal(f"Không tìm thấy file ảnh nào trong thư mục: {self.image_folder}")
            
    def _run_processing_in_thread(self, target_function, *args):
        """Chạy một hàm trong luồng riêng để không làm treo giao diện"""
        def thread_target_wrapper():
            target_function(*args)
            # Sau khi chạy xong, cập nhật lại GUI từ luồng chính
            self.master.after(0, self.gui._reset_run_button)

        threading.Thread(target=thread_target_wrapper, daemon=True).start()

    def run_single_step(self):
        """Chạy chấm cho 1 ảnh và dừng lại"""
        if not self._validate_inputs():
            self.gui._reset_run_button(); return
            
        if self.current_image_index >= len(self.image_files):
            messagebox.showinfo("Hoàn thành", "Đã chấm xong tất cả các ảnh!")
            self.gui.log_terminal("Đã chấm xong tất cả ảnh."); self.gui._reset_run_button(); return

        image_to_process = self.image_files[self.current_image_index]
        self._run_processing_in_thread(self._single_step_logic, image_to_process)
        
    def _single_step_logic(self, image_to_process):
        """Logic thực tế của việc xử lý 1 ảnh"""
        process_single_image(
            image_path=image_to_process, model_path=self.model_path,
            class_list_path=self.class_list_path, answer_key_path=self.answer_key_path,
            output_dir=self.output_dir, gui_app=self.gui
        )
        self.current_image_index += 1


    def start_auto_processing(self):
        """Bắt đầu quá trình chấm tự động"""
        if not self._validate_inputs():
            self.gui.stop_auto_mode(); return
            
        if self.current_image_index >= len(self.image_files):
            messagebox.showinfo("Hoàn thành", "Đã chấm xong tất cả các ảnh!")
            self.gui.log_terminal("Đã chấm xong tất cả ảnh."); self.gui.stop_auto_mode(); return

        self.is_auto_running = True
        threading.Thread(target=self._auto_processing_logic, daemon=True).start()

    def _auto_processing_logic(self):
        """Logic vòng lặp của chế độ tự động"""
        while self.current_image_index < len(self.image_files) and self.is_auto_running:
            image_to_process = self.image_files[self.current_image_index]
            process_single_image(
                image_path=image_to_process, model_path=self.model_path,
                class_list_path=self.class_list_path, answer_key_path=self.answer_key_path,
                output_dir=self.output_dir, gui_app=self.gui
            )
            self.current_image_index += 1
            #time.sleep(0.5) # Thêm một khoảng dừng nhỏ giữa các ảnh
        
        # Sau khi vòng lặp kết thúc
        def on_complete():
            if self.is_auto_running: # Nếu kết thúc tự nhiên (hết ảnh)
                messagebox.showinfo("Hoàn thành", f"Đã tự động chấm xong tất cả {len(self.image_files)} ảnh!")
                self.gui.log_terminal("Đã tự động chấm xong tất cả ảnh.")
            self.is_auto_running = False
            self.gui.stop_auto_mode()
        
        self.master.after(0, on_complete)

    def stop_auto_processing(self):
        """Dừng quá trình tự động"""
        self.is_auto_running = False
        self.gui.log_terminal("Quá trình tự động đã được yêu cầu dừng.")

# --- ĐIỂM KHỞI CHẠY CỦA CHƯƠNG TRÌNH ---
if __name__ == "__main__":
    root = tk.Tk()
    controller = AppController(root)
    root.mainloop()
